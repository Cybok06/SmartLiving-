from flask import Blueprint, render_template, jsonify, request, send_file
from bson import ObjectId
from datetime import datetime
from typing import Any, Dict, List
from io import BytesIO
import re

from db import db

hr_payroll_bp = Blueprint("hr_payroll", __name__, url_prefix="/hr/payroll")

users_col     = db["users"]
payrolls_col  = db["payrolls"]
settings_col  = db["payroll_settings"]   # ✅ collection for defaults
payroll_records_col    = db["payroll_records"]
payroll_deductions_col = db["payroll_deductions"]
payroll_audit_col     = db["payroll_audit_logs"]


# ---------------------------- Indexes (safe) ----------------------------
def _ensure_indexes():
    try:
        payrolls_col.create_index([("payroll_month", 1), ("status", 1), ("updated_at", -1)])
        payrolls_col.create_index([("manager_id", 1), ("payroll_month", 1)], unique=True)
        users_col.create_index([("role", 1), ("branch", 1)])

        settings_col.create_index([("scope", 1), ("branch", 1), ("month", 1)], unique=True)
        settings_col.create_index([("month", 1), ("updated_at", -1)])
        payroll_records_col.create_index([("employee_id", 1), ("month", 1)], unique=True)
        payroll_records_col.create_index([("month", 1), ("role", 1)])
        payroll_deductions_col.create_index([("employee_id", 1), ("month", 1), ("created_at", -1)])
        payroll_deductions_col.create_index([("month", 1), ("created_at", -1)])
    except Exception:
        pass

_ensure_indexes()


# ---------------------------- Helpers ----------------------------
EDITABLE_HR_STATUSES = {"Submitted", "Under Review"}
DECISION_STATUSES   = {"Submitted", "Under Review"}

WATERMARK_URL = "https://imagedelivery.net/h9fmMoa1o2c2P55TcWJGOg/8744779c-1300-4a50-49de-b143d24da300/public"

DEFAULT_TAX_CODE = "default_tax"  # ✅ internal identifier for the tax item in global_deductions

def _now():
    return datetime.utcnow()

def _audit_payroll(action: str, entity: str, entity_id: str = "", meta: Dict[str, Any] | None = None):
    try:
        payroll_audit_col.insert_one({
            "action": action,
            "entity": entity,
            "entity_id": entity_id,
            "meta": meta or {},
            "created_at": _now(),
        })
    except Exception:
        pass

def _oid(val):
    try:
        return ObjectId(val)
    except Exception:
        return None

def _normalize_month(month: str) -> str:
    if not month or len(month) != 7 or month[4] != "-":
        raise ValueError("Invalid month format. Use YYYY-MM.")
    y = int(month[:4])
    m = int(month[5:7])
    if m < 1 or m > 12:
        raise ValueError("Invalid month.")
    if y < 2000 or y > 2100:
        raise ValueError("Invalid year.")
    return month

def _month_label(yyyy_mm: str) -> str:
    try:
        y, m = yyyy_mm.split("-")
        m = int(m)
        months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
        return f"{months[m-1]} {y}"
    except Exception:
        return yyyy_mm

def _to_float(x, default=0.0):
    try:
        if x is None or x == "":
            return default
        return float(x)
    except Exception:
        return default

def _money(n: float) -> float:
    try:
        return round(float(n or 0.0), 2)
    except Exception:
        return 0.0

def _clean_spaces(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s

def _norm_branch(branch: str) -> str:
    b = _clean_spaces(branch)
    low = b.lower()

    hq_syn = {
        "hq", "headquarter", "headquarters",
        "head quarter", "head quarters",
        "head quater", "head quaters",
        "headquater", "headquaters",
        "headquater ", "headquarters "
    }
    low2 = low.replace(".", "").strip()
    low2 = re.sub(r"\s+", " ", low2)

    if low2 in hq_syn:
        return "HQ"

    if len(b) <= 4 and b.isupper():
        return b
    return b.title() if b else "Unknown"

def _manager_maps():
    managers = list(users_col.find(
        {"role": "manager"},
        {"_id": 1, "name": 1, "branch": 1, "phone": 1, "email": 1, "image_url": 1}
    ))
    by_str = {str(m["_id"]): m for m in managers}
    by_oid = {m["_id"]: m for m in managers}
    return by_str, by_oid

def _get_manager_from_payroll(mid, by_str, by_oid):
    if isinstance(mid, ObjectId):
        return by_oid.get(mid) or {}
    if isinstance(mid, str):
        return by_str.get(mid) or {}
    return {}

def _final_gross_agent(item: dict) -> float:
    ma = _to_float(item.get("manager_amount"), 0.0)
    hr_amt = item.get("hr_amount", None)
    if hr_amt is None or hr_amt == "":
        return max(0.0, ma)
    return max(0.0, _to_float(hr_amt, 0.0))

def _apply_global_deductions(gross: float, global_deds) -> float:
    gross = max(0.0, _to_float(gross, 0.0))
    total = 0.0
    for d in (global_deds or []):
        dtype = (d.get("type") or "fixed").lower().strip()
        val = _to_float(d.get("value"), 0.0)
        if val < 0:
            val = 0.0
        if dtype == "percent":
            total += (gross * (val / 100.0))
        else:
            total += val
    return total

def _sum_deductions(deds) -> float:
    total = 0.0
    for d in (deds or []):
        amt = _to_float(d.get("amount"), 0.0)
        if amt < 0:
            amt = 0.0
        total += amt
    return total

def _load_deductions(employee_id: str, month: str) -> List[Dict[str, Any]]:
    cursor = payroll_deductions_col.find(
        {"employee_id": employee_id, "month": month},
        {"amount": 1, "reason": 1, "date": 1},
    ).sort("created_at", 1)
    rows = []
    for d in cursor:
        rows.append({
            "deduction_id": str(d.get("_id")),
            "amount": _money(_to_float(d.get("amount"), 0.0)),
            "reason": d.get("reason") or "",
            "date": (d.get("date") or ""),
        })
    return rows

def _ensure_payroll_record_exists(employee_id: str, month: str, user_fields: Dict[str, Any] | None = None) -> Dict[str, Any]:
    record = payroll_records_col.find_one({"employee_id": employee_id, "month": month})
    if record:
        return record

    uf = user_fields or {}
    now = _now()
    doc = {
        "employee_id": employee_id,
        "employee_name": uf.get("employee_name") or "",
        "role": uf.get("role") or "",
        "branch": uf.get("branch") or "",
        "month": month,
        "basic_salary": 0.0,
        "allowances": 0.0,
        "gross": 0.0,
        "ssf_employee": 0.0,
        "paye": 0.0,
        "deductions_total": 0.0,
        "net_pay": 0.0,
        "employer_13": 0.0,
        "total_staff_cost": 0.0,
        "tier1": 0.0,
        "tier2": 0.0,
        "deductions": [],
        "created_by": uf.get("created_by") or "hr",
        "created_at": now,
        "updated_at": now,
    }
    payroll_records_col.insert_one(doc)
    _audit_payroll("payroll_created", "payroll_record", entity_id=f"{employee_id}:{month}")
    return doc

def _recompute_record(employee_id: str, month: str) -> Dict[str, Any] | None:
    rec = payroll_records_col.find_one({"employee_id": employee_id, "month": month})
    if not rec:
        return None

    deductions = _load_deductions(employee_id, month)
    basic = _money(_to_float(rec.get("basic_salary"), 0.0))
    allowances = _money(_to_float(rec.get("allowances"), 0.0))
    gross = _money(basic + allowances)
    ssf_employee = _money(basic * 0.055)
    paye = _money(_to_float(rec.get("paye"), 0.0))
    deductions_total = _money(sum(_to_float(d.get("amount"), 0.0) for d in deductions))
    net_pay = _money(gross - ssf_employee - paye - deductions_total)
    employer_13 = _money(basic * 0.13)
    total_staff_cost = _money(net_pay + employer_13)
    tier1 = _money(basic * 0.135)
    tier2 = _money(basic * 0.05)

    update = {
        "basic_salary": basic,
        "allowances": allowances,
        "gross": gross,
        "ssf_employee": ssf_employee,
        "paye": paye,
        "deductions": deductions,
        "deductions_total": deductions_total,
        "net_pay": net_pay,
        "employer_13": employer_13,
        "total_staff_cost": total_staff_cost,
        "tier1": tier1,
        "tier2": tier2,
        "updated_at": _now(),
    }
    payroll_records_col.update_one({"_id": rec["_id"]}, {"$set": update})
    rec.update(update)
    return rec

def _apply_deductions_to_record(record: Dict[str, Any], deductions: List[Dict[str, Any]]) -> Dict[str, Any]:
    deductions_total = _money(sum(_to_float(d.get("amount"), 0.0) for d in deductions))
    gross = _to_float(record.get("gross"), 0.0)
    ssf_employee = _to_float(record.get("ssf_employee"), 0.0)
    paye = _to_float(record.get("paye"), 0.0)
    net_pay = _money(gross - ssf_employee - paye - deductions_total)
    employer_13 = _to_float(record.get("employer_13"), 0.0)
    total_staff_cost = _money(net_pay + employer_13)

    record["deductions"] = deductions
    record["deductions_total"] = deductions_total
    record["net_pay"] = net_pay
    record["total_staff_cost"] = total_staff_cost
    record["updated_at"] = _now()
    return record

# ✅ NEW: adjustments (add/deduct) summation
def _sum_adjustments(adjs):
    add_total = 0.0
    deduct_total = 0.0
    for a in (adjs or []):
        atype = (a.get("type") or "").strip().lower()
        amt = _to_float(a.get("amount"), 0.0)
        if amt < 0:
            amt = 0.0
        if atype == "add":
            add_total += amt
        elif atype == "deduct":
            deduct_total += amt
    return (_money(add_total), _money(deduct_total))

def _net_agent(item: dict, global_deds) -> float:
    gross = _final_gross_agent(item)
    item_deds = _sum_deductions(item.get("deductions"))
    glob_deds = _apply_global_deductions(gross, global_deds)

    # ✅ include adjustments in net
    add_total, deduct_total = _sum_adjustments(item.get("adjustments"))
    net = gross - item_deds - glob_deds - deduct_total + add_total
    return max(0.0, _money(net))

def _manager_pay_final(payroll_doc: dict) -> float:
    mp = payroll_doc.get("manager_pay") or {}
    amt = mp.get("amount", None)
    if amt is None or amt == "":
        return 0.0
    return max(0.0, _to_float(amt, 0.0))

def _manager_pay_net(payroll_doc: dict) -> float:
    global_deds = payroll_doc.get("global_deductions") or []
    gross = _manager_pay_final(payroll_doc)
    return max(0.0, _money(gross - _apply_global_deductions(gross, global_deds)))

def _recompute_totals(payroll_doc: dict):
    items = (payroll_doc.get("items") or [])
    global_deds = payroll_doc.get("global_deductions") or []

    submitted_total = 0.0
    gross_final_total = 0.0
    net_final_total = 0.0
    edited_count = 0

    for it in items:
        submitted_total += _to_float(it.get("manager_amount"), 0.0)
        gross_final_total += _final_gross_agent(it)
        net_final_total += _net_agent(it, global_deds)
        if it.get("changed_by_hr"):
            edited_count += 1

    mp_gross = _manager_pay_final(payroll_doc)
    mp_net   = _manager_pay_net(payroll_doc)
    gross_final_total += mp_gross
    net_final_total   += mp_net

    return {
        "submitted_total": _money(submitted_total),
        "final_total": _money(gross_final_total),
        "net_final_total": _money(net_final_total),
        "edited_count": int(edited_count),
        "agent_count": int(len(items)),
        "manager_pay_gross": _money(mp_gross),
        "manager_pay_net": _money(mp_net),
    }

def _compute_stats(items, global_deds=None):
    items = items or []
    global_deds = global_deds or []
    if not items:
        return {
            "count": 0,
            "avg_net": 0.0,
            "highest_net": 0.0,
            "highest_agent": None,
            "lowest_net": 0.0,
            "lowest_agent": None,
        }

    nets = [(float(_net_agent(it, global_deds)), it) for it in items]
    total = sum(v for v, _ in nets)
    avg = total / max(len(nets), 1)

    highest = max(nets, key=lambda x: x[0])
    lowest  = min(nets, key=lambda x: x[0])

    return {
        "count": len(nets),
        "avg_net": _money(avg),
        "highest_net": _money(highest[0]),
        "highest_agent": {
            "agent_id": str(highest[1].get("agent_id") or ""),
            "agent_name": highest[1].get("agent_name") or "Agent",
            "agent_branch": _norm_branch(highest[1].get("agent_branch") or ""),
        },
        "lowest_net": _money(lowest[0]),
        "lowest_agent": {
            "agent_id": str(lowest[1].get("agent_id") or ""),
            "agent_name": lowest[1].get("agent_name") or "Agent",
            "agent_branch": _norm_branch(lowest[1].get("agent_branch") or ""),
        },
    }

def _infer_branch_from_items(items):
    counter = {}
    for it in (items or []):
        b = _norm_branch(it.get("agent_branch") or "")
        if not b:
            continue
        key = b.lower()
        counter[key] = counter.get(key, 0) + 1
    if not counter:
        return "Unknown"
    best = max(counter.items(), key=lambda x: x[1])[0]
    return _norm_branch(best)

def _overview_row(row_type: str, employee_id: str, name: str, role: str, gross: float,
                  global_deds_total: float, other_deds_total: float,
                  adj_add: float, adj_deduct: float, net: float,
                  note: str, editable: bool, payroll_id: str | None = None,
                  extra: Dict[str, Any] | None = None):
    row = {
        "row_type": row_type,
        "employee_id": employee_id,
        "name": name,
        "role": role,
        "gross": _money(gross),
        "global_deductions_total": _money(global_deds_total),
        "deductions_total": _money(other_deds_total),
        "adjustments_add": _money(adj_add),
        "adjustments_deduct": _money(adj_deduct),
        "net": _money(net),
        "note": (note or "").strip(),
        "editable": bool(editable),
        "payroll_id": payroll_id or "",
    }
    if extra:
        row.update(extra)
    return row


# ---------------------------- Defaults: TAX helpers ----------------------------
def _clean_tax_payload(tax: dict):
    """
    tax = {name, type: fixed|percent, value}
    """
    tax = tax or {}
    name = (tax.get("name") or "Tax").strip()
    dtype = (tax.get("type") or "percent").lower().strip()
    if dtype not in {"fixed", "percent"}:
        dtype = "percent"
    value = _to_float(tax.get("value"), 0.0)
    if value < 0:
        value = 0.0
    return {
        "name": name,
        "type": dtype,
        "value": _money(value),
        "code": DEFAULT_TAX_CODE,     # ✅ so we can reliably update/replace it
    }

def _upsert_global_deduction(global_deds: list, item: dict):
    """
    Replace any existing item with same code, else append.
    """
    global_deds = list(global_deds or [])
    code = (item.get("code") or "").strip()
    if not code:
        global_deds.append(item)
        return global_deds

    replaced = False
    for i, d in enumerate(global_deds):
        if (d.get("code") or "").strip() == code:
            global_deds[i] = item
            replaced = True
            break
    if not replaced:
        global_deds.append(item)
    return global_deds

def _get_tax_default(month: str, branch: str | None):
    """
    Returns (effective_tax, all_tax, branch_tax)
    month-specific first, else month=None fallback.
    """
    b = _norm_branch(branch or "") if branch else None

    # 1) branch month
    branch_tax = None
    if b:
        branch_tax = settings_col.find_one({"scope": "branch", "branch": b, "month": month})
        if not branch_tax:
            branch_tax = settings_col.find_one({"scope": "branch", "branch": b, "month": None})

    # 2) all month
    all_tax = settings_col.find_one({"scope": "all", "branch": None, "month": month})
    if not all_tax:
        all_tax = settings_col.find_one({"scope": "all", "branch": None, "month": None})

    eff = (branch_tax or all_tax or {})
    return (eff.get("tax") if eff else None), (all_tax.get("tax") if all_tax else None), (branch_tax.get("tax") if branch_tax else None)

def _ensure_default_tax_on_payroll(pdoc: dict, effective_tax: dict | None):
    """
    If payroll is editable and default tax exists, inject it when missing.
    """
    if not pdoc:
        return False

    status = pdoc.get("status", "Draft")
    if status not in EDITABLE_HR_STATUSES:
        return False

    if not effective_tax:
        return False

    g = pdoc.get("global_deductions") or []
    has = any((d.get("code") or "") == DEFAULT_TAX_CODE for d in g)
    if has:
        return False

    g2 = _upsert_global_deduction(g, effective_tax)
    now = _now()
    payrolls_col.update_one(
        {"_id": pdoc["_id"]},
        {"$set": {
            "global_deductions": g2,
            "updated_at": now,
            "status": "Under Review" if status == "Submitted" else status,
        }}
    )

    p2 = payrolls_col.find_one({"_id": pdoc["_id"]})
    totals = _recompute_totals(p2 or {})
    payrolls_col.update_one({"_id": pdoc["_id"]}, {"$set": {"totals": totals, "updated_at": now}})
    return True


# ---------------------------- NEW: Payroll deductions (HR controlled) ----------------------------
@hr_payroll_bp.route("/deductions/add", methods=["POST"])
def hr_add_payroll_deduction():
    data = request.get_json(silent=True) or {}
    employee_id = str(data.get("employee_id") or "").strip()
    amount = data.get("amount")
    reason = (data.get("reason") or "").strip()
    date_str = (data.get("date") or "").strip()
    month = (data.get("month") or "").strip()

    if not employee_id or not month:
        return jsonify(ok=False, message="Missing employee_id or month."), 400
    if not reason:
        return jsonify(ok=False, message="Reason is required."), 400
    if not date_str:
        return jsonify(ok=False, message="Date is required."), 400
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        return jsonify(ok=False, message="Invalid date. Use YYYY-MM-DD."), 400
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400
    try:
        amt = float(amount)
    except Exception:
        return jsonify(ok=False, message="Invalid amount."), 400
    if amt <= 0:
        return jsonify(ok=False, message="Amount must be greater than 0."), 400

    user = users_col.find_one({"_id": _oid(employee_id)}) if _oid(employee_id) else None
    if not user:
        user = users_col.find_one({"_id": employee_id})

    _ensure_payroll_record_exists(employee_id, month, {
        "employee_name": (user.get("name") if user else ""),
        "role": (user.get("role") if user else ""),
        "branch": (user.get("branch") if user else ""),
        "created_by": (data.get("created_by") or "hr"),
    })

    now = _now()
    doc = {
        "employee_id": employee_id,
        "amount": _money(amt),
        "reason": reason,
        "date": date_str,
        "month": month,
        "created_by": (data.get("created_by") or "hr"),
        "created_at": now,
    }
    res = payroll_deductions_col.insert_one(doc)

    _recompute_record(employee_id, month)

    _audit_payroll("deduction_added", "payroll_deduction", entity_id=str(res.inserted_id), meta={
        "employee_id": employee_id,
        "month": month,
        "amount": _money(amt),
    })

    return jsonify(ok=True, message="Deduction added.", deduction_id=str(res.inserted_id))


@hr_payroll_bp.route("/deductions", methods=["GET"])
def hr_list_payroll_deductions():
    employee_id = (request.args.get("employee_id") or "").strip()
    month = (request.args.get("month") or "").strip()

    if not employee_id or not month:
        return jsonify(ok=False, message="Missing employee_id or month."), 400
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    deductions = _load_deductions(employee_id, month)
    return jsonify(ok=True, employee_id=employee_id, month=month, deductions=deductions), 200


@hr_payroll_bp.route("/deductions/bulk", methods=["POST"])
def hr_add_payroll_deduction_bulk():
    data = request.get_json(silent=True) or {}
    employee_ids = data.get("employee_ids") or []
    apply_all = bool(data.get("apply_all"))
    amount = data.get("amount")
    reason = (data.get("reason") or "").strip()
    date_str = (data.get("date") or "").strip()
    month = (data.get("month") or "").strip()

    if not month:
        return jsonify(ok=False, message="Missing month (YYYY-MM)."), 400
    if not reason:
        return jsonify(ok=False, message="Reason is required."), 400
    if not date_str:
        return jsonify(ok=False, message="Date is required."), 400
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
    except Exception:
        return jsonify(ok=False, message="Invalid date. Use YYYY-MM-DD."), 400
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400
    try:
        amt = float(amount)
    except Exception:
        return jsonify(ok=False, message="Invalid amount."), 400
    if amt <= 0:
        return jsonify(ok=False, message="Amount must be greater than 0."), 400

    if apply_all:
        employees = list(users_col.find({}, {"_id": 1, "name": 1, "role": 1, "branch": 1}))
    else:
        if not isinstance(employee_ids, list) or not employee_ids:
            return jsonify(ok=False, message="Select at least one employee."), 400
        employees = list(users_col.find(
            {"_id": {"$in": [ObjectId(x) for x in employee_ids if _oid(x)]}},
            {"_id": 1, "name": 1, "role": 1, "branch": 1}
        ))

    if not employees:
        return jsonify(ok=False, message="No employees found to apply deduction."), 404

    now = _now()
    docs = []
    for u in employees:
        emp_id = str(u.get("_id"))
        _ensure_payroll_record_exists(emp_id, month, {
            "employee_name": u.get("name") or "",
            "role": u.get("role") or "",
            "branch": u.get("branch") or "",
            "created_by": (data.get("created_by") or "hr"),
        })
        docs.append({
            "employee_id": emp_id,
            "amount": _money(amt),
            "reason": reason,
            "date": date_str,
            "month": month,
            "created_by": (data.get("created_by") or "hr"),
            "created_at": now,
        })

    if docs:
        payroll_deductions_col.insert_many(docs)
        for d in docs:
            _recompute_record(d["employee_id"], month)

    _audit_payroll("deduction_added_bulk", "payroll_deduction", meta={
        "month": month,
        "amount": _money(amt),
        "reason": reason,
        "count": len(docs),
        "apply_all": apply_all,
    })

    return jsonify(ok=True, message="Deduction applied.", applied=len(docs)), 200


@hr_payroll_bp.route("/records/upsert", methods=["POST"])
def hr_upsert_payroll_record():
    data = request.get_json(silent=True) or {}
    employee_id = str(data.get("employee_id") or "").strip()
    month = (data.get("month") or "").strip()

    if not employee_id or not month:
        return jsonify(ok=False, message="Missing employee_id or month."), 400
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    existing = payroll_records_col.find_one({"employee_id": employee_id, "month": month})

    emp = users_col.find_one({"_id": _oid(employee_id)}) if _oid(employee_id) else None
    if not emp:
        emp = users_col.find_one({"_id": employee_id})

    basic = _money(_to_float(data.get("basic_salary"), 0.0))
    allowances = _money(_to_float(data.get("allowances"), 0.0))
    gross = _money(basic + allowances)
    ssf_employee = _money(basic * 0.055)
    taxable = _money(gross - ssf_employee)
    paye = _money(_to_float(data.get("paye"), 0.0))
    employer_13 = _money(basic * 0.13)
    tier1 = _money(basic * 0.135)
    tier2 = _money(basic * 0.05)

    deductions = _load_deductions(employee_id, month)
    deductions_total = _money(sum(_to_float(d.get("amount"), 0.0) for d in deductions))
    net_pay = _money(gross - ssf_employee - paye - deductions_total)
    total_staff_cost = _money(net_pay + employer_13)

    now = _now()
    status_val = data.get("status")
    if status_val is None:
        status_val = (existing.get("status") if existing else "") or ""
    note_val = data.get("note")
    if note_val is None:
        note_val = (existing.get("note") if existing else "") or ""

    rec = {
        "employee_id": employee_id,
        "employee_name": data.get("employee_name") or (emp.get("name") if emp else ""),
        "role": data.get("role") or (emp.get("role") if emp else ""),
        "branch": data.get("branch") or (emp.get("branch") if emp else ""),
        "month": month,
        "status": str(status_val).strip(),
        "note": str(note_val).strip(),
        "basic_salary": basic,
        "allowances": allowances,
        "gross": gross,
        "ssf_employee": ssf_employee,
        "paye": paye,
        "deductions_total": deductions_total,
        "net_pay": net_pay,
        "employer_13": employer_13,
        "total_staff_cost": total_staff_cost,
        "tier1": tier1,
        "tier2": tier2,
        "deductions": deductions,
        "created_by": data.get("created_by") or "hr",
        "updated_at": now,
    }

    payroll_records_col.update_one(
        {"employee_id": employee_id, "month": month},
        {"$set": rec, "$setOnInsert": {"created_at": now}},
        upsert=True,
    )

    _audit_payroll("payroll_created", "payroll_record", entity_id=f"{employee_id}:{month}")
    return jsonify(ok=True, message="Payroll record saved.", employee_id=employee_id, month=month)


@hr_payroll_bp.route("/records", methods=["GET"])
def hr_payroll_records_list():
    month = (request.args.get("month") or "").strip()
    role = (request.args.get("role") or "").strip()
    branch = (request.args.get("branch") or "").strip()
    status = (request.args.get("status") or "").strip()
    search = (request.args.get("search") or "").strip()

    if not month:
        return jsonify(ok=False, message="Missing month (YYYY-MM)."), 400
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    user_q: Dict[str, Any] = {}
    if role:
        user_q["role"] = {"$regex": f"^{re.escape(role)}$", "$options": "i"}
    if branch:
        user_q["branch"] = {"$regex": f"^{re.escape(branch)}$", "$options": "i"}
    if search:
        pattern = re.escape(search)
        user_q["$or"] = [
            {"name": {"$regex": pattern, "$options": "i"}},
            {"role": {"$regex": pattern, "$options": "i"}},
            {"branch": {"$regex": pattern, "$options": "i"}},
            {"username": {"$regex": pattern, "$options": "i"}},
        ]

    users = list(users_col.find(user_q, {"_id": 1, "name": 1, "role": 1, "branch": 1}))
    records = list(payroll_records_col.find({"month": month}))
    rec_by_emp = {str(r.get("employee_id")): r for r in records}
    user_ids = {str(u.get("_id")) for u in users}

    rows = []
    for u in users:
        uid = str(u.get("_id"))
        rec = rec_by_emp.get(uid)
        if status and status != "All":
            if not rec or (rec.get("status") or "") != status:
                continue
        if rec:
            rec_out = dict(rec)
            rec_out["_id"] = str(rec.get("_id"))
            rows.append(rec_out)
        else:
            rows.append({
                "_id": "",
                "employee_id": uid,
                "employee_name": u.get("name") or "",
                "role": u.get("role") or "",
                "branch": u.get("branch") or "",
                "month": month,
                "basic_salary": 0.0,
                "allowances": 0.0,
                "gross": 0.0,
                "ssf_employee": 0.0,
                "paye": 0.0,
                "deductions_total": 0.0,
                "net_pay": 0.0,
                "employer_13": 0.0,
                "total_staff_cost": 0.0,
                "tier1": 0.0,
                "tier2": 0.0,
                "deductions": [],
                "status": "",
            })

    record_only = []
    for r in records:
        if str(r.get("employee_id")) not in user_ids:
            if search:
                s = search.lower()
                name = (r.get("employee_name") or "").lower()
                role2 = (r.get("role") or "").lower()
                branch2 = (r.get("branch") or "").lower()
                if s not in name and s not in role2 and s not in branch2:
                    continue
            if status and status != "All":
                if (r.get("status") or "") != status:
                    continue
            ro = dict(r)
            ro["_id"] = str(r.get("_id"))
            record_only.append(ro)

    rows.extend(record_only)
    rows.sort(key=lambda x: (x.get("employee_name") or "").lower())
    return jsonify(ok=True, month=month, rows=rows)


@hr_payroll_bp.route("/records/employee/<employee_id>", methods=["GET"])
def hr_payroll_record_employee(employee_id):
    month = (request.args.get("month") or "").strip()
    if not month:
        return jsonify(ok=False, message="Missing month (YYYY-MM)."), 400
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    rec = payroll_records_col.find_one({"employee_id": employee_id, "month": month})
    if not rec:
        user = users_col.find_one({"_id": _oid(employee_id)}) if _oid(employee_id) else None
        if not user:
            user = users_col.find_one({"_id": employee_id})
        if not user:
            return jsonify(ok=True, found=False, message="No payroll record found.")
        rec = _ensure_payroll_record_exists(employee_id, month, {
            "employee_name": user.get("name") or "",
            "role": user.get("role") or "",
            "branch": user.get("branch") or "",
            "created_by": "hr",
        })
        rec = payroll_records_col.find_one({"employee_id": employee_id, "month": month})
    rec = _recompute_record(employee_id, month) or rec
    rec["_id"] = str(rec.get("_id"))
    return jsonify(ok=True, found=True, record=rec)


@hr_payroll_bp.route("/employees", methods=["GET"])
def hr_payroll_employees():
    q = (request.args.get("q") or "").strip()
    query: Dict[str, Any] = {}
    if q:
        pattern = re.escape(q)
        query["$or"] = [
            {"name": {"$regex": pattern, "$options": "i"}},
            {"role": {"$regex": pattern, "$options": "i"}},
            {"branch": {"$regex": pattern, "$options": "i"}},
            {"username": {"$regex": pattern, "$options": "i"}},
        ]

    cur = users_col.find(query, {"name": 1, "role": 1, "branch": 1}).sort("name", 1).limit(500)
    rows = []
    for u in cur:
        rows.append({
            "id": str(u.get("_id")),
            "name": u.get("name") or "",
            "role": u.get("role") or "",
            "branch": u.get("branch") or "",
        })
    return jsonify(ok=True, employees=rows), 200
# ---------------------------- Views ----------------------------
@hr_payroll_bp.route("/", methods=["GET"], strict_slashes=False)
def hr_payroll_home():
    return render_template(
        "hr_pages/hr_payroll.html",
        hr_name="HR",
        today_iso=_now().date().isoformat(),
        active_page="payroll"
    )


# ---------------------------- JSON: month options ----------------------------
@hr_payroll_bp.route("/months", methods=["GET"])
def hr_payroll_months():
    months = payrolls_col.distinct("payroll_month")
    months = sorted([m for m in months if isinstance(m, str)], reverse=True)
    if not months:
        now = _now()
        months = [f"{now.year}-{str(now.month).zfill(2)}"]

    return jsonify(ok=True, months=[{"value": m, "label": _month_label(m)} for m in months]), 200


# ---------------------------- JSON: Get/Set tax defaults ----------------------------
@hr_payroll_bp.route("/tax-defaults", methods=["GET"])
def hr_get_tax_defaults():
    month  = request.args.get("month") or ""
    branch = request.args.get("branch") or ""
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    eff, all_tax, branch_tax = _get_tax_default(month, branch if branch else None)
    return jsonify(ok=True, month=month, branch=_norm_branch(branch) if branch else None,
                   effective=eff, all_default=all_tax, branch_default=branch_tax), 200


@hr_payroll_bp.route("/tax-defaults", methods=["POST"])
def hr_set_tax_defaults():
    """
    payload:
    {
      month: "YYYY-MM" (required),
      scope: "all" | "branch",
      branch: "Kasoa" (required if scope=branch),
      tax: {name,type,value},
      apply: true|false   # apply immediately to editable payrolls in that scope+month
    }
    """
    data = request.get_json(silent=True) or {}
    month  = data.get("month") or ""
    scope  = (data.get("scope") or "all").strip().lower()
    branch = data.get("branch") or None
    apply_ = bool(data.get("apply", True))

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    if scope not in {"all", "branch"}:
        return jsonify(ok=False, message="Invalid scope. Use 'all' or 'branch'."), 400

    if scope == "branch":
        if not branch:
            return jsonify(ok=False, message="Branch is required for branch scope."), 400
        branch = _norm_branch(branch)
    else:
        branch = None

    tax_item = _clean_tax_payload(data.get("tax") or {})
    now = _now()

    settings_col.update_one(
        {"scope": scope, "branch": branch, "month": month},
        {"$set": {
            "scope": scope,
            "branch": branch,
            "month": month,
            "tax": tax_item,
            "updated_at": now
        }},
        upsert=True
    )

    applied = 0
    if apply_:
        applied = _apply_tax_default_to_payrolls(month=month, scope=scope, branch=branch, tax_item=tax_item)

    return jsonify(ok=True, message="Tax default saved.", applied=applied, tax=tax_item), 200


def _apply_tax_default_to_payrolls(month: str, scope: str, branch: str | None, tax_item: dict):
    """
    Apply default tax to payrolls in a month:
      - editable statuses only
      - scope=all => all branches
      - scope=branch => inferred branch match
    """
    q = {"payroll_month": month, "status": {"$in": list(EDITABLE_HR_STATUSES)}}
    cur = list(payrolls_col.find(q, {"items": 1, "manager_id": 1, "global_deductions": 1, "status": 1, "totals": 1, "manager_pay": 1}))

    mgr_by_str, mgr_by_oid = _manager_maps()

    count = 0
    for p in cur:
        items = p.get("items", []) or []
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)

        branch_raw = (mgr.get("branch") or "").strip()
        inferred = _norm_branch(branch_raw) if branch_raw else _infer_branch_from_items(items)
        inferred = _norm_branch(inferred)

        if scope == "branch":
            if _norm_branch(branch or "") != inferred:
                any_match = any(_norm_branch(it.get("agent_branch") or "") == _norm_branch(branch or "") for it in items)
                if not any_match:
                    continue

        g = p.get("global_deductions") or []
        g2 = _upsert_global_deduction(g, tax_item)

        now = _now()
        new_status = "Under Review" if (p.get("status") == "Submitted") else p.get("status")

        payrolls_col.update_one(
            {"_id": p["_id"]},
            {"$set": {"global_deductions": g2, "updated_at": now, "status": new_status}}
        )

        p2 = payrolls_col.find_one({"_id": p["_id"]})
        totals = _recompute_totals(p2 or {})
        payrolls_col.update_one({"_id": p["_id"]}, {"$set": {"totals": totals, "updated_at": now}})
        count += 1

    return count


# ---------------------------- JSON: branch summary (normalized) ----------------------------
@hr_payroll_bp.route("/branch-summary", methods=["GET"])
def hr_branch_summary():
    month  = request.args.get("month") or ""
    status = request.args.get("status") or "Submitted"

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    mgr_by_str, mgr_by_oid = _manager_maps()

    q = {"payroll_month": month}
    if status and status != "All":
        q["status"] = status

    cur = payrolls_col.find(q, {"items": 1, "manager_id": 1, "totals": 1, "global_deductions": 1, "manager_pay": 1})

    bucket = {}

    for p in cur:
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)
        items = p.get("items", []) or []

        branch_raw = (mgr.get("branch") or "").strip()
        branch = _norm_branch(branch_raw) if branch_raw else _infer_branch_from_items(items)

        b = bucket.setdefault(branch, {
            "branch": branch,
            "payroll_count": 0,
            "agent_count": 0,
            "submitted_total": 0.0,
            "final_total": 0.0,
            "net_final_total": 0.0,
            "edited_count": 0,
            "manager_pay_gross": 0.0,
        })

        b["payroll_count"] += 1
        b["agent_count"] += len(items)

        totals = p.get("totals") or _recompute_totals(p)

        b["submitted_total"] += _to_float(totals.get("submitted_total"), 0)
        b["final_total"]     += _to_float(totals.get("final_total"), 0)
        b["net_final_total"] += _to_float(totals.get("net_final_total"), 0)
        b["edited_count"]    += int(totals.get("edited_count") or 0)
        b["manager_pay_gross"] += _to_float(totals.get("manager_pay_gross"), 0)

    rows = sorted(bucket.values(), key=lambda x: str(x["branch"]).lower())
    for r in rows:
        r["submitted_total"] = _money(r["submitted_total"])
        r["final_total"]     = _money(r["final_total"])
        r["net_final_total"] = _money(r["net_final_total"])
        r["manager_pay_gross"] = _money(r["manager_pay_gross"])

    return jsonify(ok=True, month=month, month_label=_month_label(month), rows=rows), 200


# ---------------------------- JSON: payroll list for a branch (normalized) ----------------------------
@hr_payroll_bp.route("/branch/<branch>/payrolls", methods=["GET"])
def hr_branch_payrolls(branch):
    month  = request.args.get("month") or ""
    status = request.args.get("status") or "Submitted"

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    branch_key = _norm_branch(branch)

    q = {"payroll_month": month}
    if status and status != "All":
        q["status"] = status

    mgr_by_str, mgr_by_oid = _manager_maps()
    cur = payrolls_col.find(
        q,
        {"items": 1, "manager_id": 1, "status": 1, "submitted_at": 1, "updated_at": 1, "totals": 1, "global_deductions": 1, "manager_pay": 1}
    ).sort("updated_at", -1).limit(300)

    rows = []
    for p in cur:
        items = p.get("items", []) or []
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)

        branch_raw = (mgr.get("branch") or "").strip()
        inferred = _norm_branch(branch_raw) if branch_raw else _infer_branch_from_items(items)

        if _norm_branch(inferred) != branch_key:
            any_match = any(_norm_branch(it.get("agent_branch") or "") == branch_key for it in items)
            if not any_match:
                continue

        totals = p.get("totals") or _recompute_totals(p)

        gds = p.get("global_deductions") or []
        has_tax = any((d.get("code") or "") == DEFAULT_TAX_CODE for d in gds)

        rows.append({
            "payroll_id": str(p["_id"]),
            "manager_id": str(mid) if mid is not None else "",
            "manager_name": (mgr.get("name") or "Manager"),
            "branch": inferred,
            "status": p.get("status", "Draft"),
            "submitted_at": p.get("submitted_at").isoformat() if p.get("submitted_at") else None,
            "updated_at": p.get("updated_at").isoformat() if p.get("updated_at") else None,
            "submitted_total": float(totals.get("submitted_total") or 0),
            "final_total": float(totals.get("final_total") or 0),
            "net_final_total": float(totals.get("net_final_total") or 0),
            "manager_pay_gross": float(totals.get("manager_pay_gross") or 0),
            "edited_count": int(totals.get("edited_count") or 0),
            "agent_count": int(totals.get("agent_count") or len(items)),
            "has_default_tax": bool(has_tax),
        })

    return jsonify(ok=True, rows=rows), 200


# ---------------------------- JSON: overview (all branches) ----------------------------
@hr_payroll_bp.route("/overview", methods=["GET"])
def hr_overview_all():
    month  = request.args.get("month") or ""
    status = request.args.get("status") or "Submitted"

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    q = {"payroll_month": month}
    if status and status != "All":
        q["status"] = status

    mgr_by_str, mgr_by_oid = _manager_maps()
    cur = payrolls_col.find(q, {"items": 1, "manager_id": 1, "status": 1, "global_deductions": 1, "manager_pay": 1})

    rows: List[Dict[str, Any]] = []
    used_ids = set()

    submitted_total = 0.0
    gross_total = 0.0
    net_total = 0.0
    edited_count = 0
    has_tax = False

    for p in cur:
        items = p.get("items", []) or []
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)

        global_deds = p.get("global_deductions") or []
        if any((d.get("code") or "") == DEFAULT_TAX_CODE for d in global_deds):
            has_tax = True

        editable = p.get("status") in EDITABLE_HR_STATUSES

        for it in items:
            agent_id = str(it.get("agent_id") or "")
            if not agent_id:
                continue

            used_ids.add(agent_id)
            ma = _to_float(it.get("manager_amount"), 0.0)
            submitted_total += ma
            if it.get("changed_by_hr"):
                edited_count += 1

            it2 = dict(it)
            it2["deductions"] = _load_deductions(agent_id, month)
            it2["adjustments"] = it2.get("adjustments") or []

            gross = _final_gross_agent(it2)
            net = _net_agent(it2, global_deds)
            global_deds_total = _apply_global_deductions(gross, global_deds)
            other_deds_total = _sum_deductions(it2.get("deductions"))
            adj_add, adj_deduct = _sum_adjustments(it2.get("adjustments"))

            gross_total += gross
            net_total += net

            rows.append(_overview_row(
                "agent",
                agent_id,
                it2.get("agent_name") or "Agent",
                "Agent",
                gross,
                global_deds_total,
                other_deds_total,
                adj_add,
                adj_deduct,
                net,
                it2.get("hr_note") or "",
                editable,
                str(p.get("_id")),
                {
                    "hr_amount": it2.get("hr_amount"),
                    "hr_note": it2.get("hr_note") or "",
                }
            ))

        mp_gross = _manager_pay_final(p)
        if mp_gross > 0:
            manager_id = str(p.get("manager_id") or "")
            used_ids.add(manager_id)

            net = _manager_pay_net(p)
            gross_total += mp_gross
            net_total += net

            rows.append(_overview_row(
                "manager",
                manager_id,
                mgr.get("name") if mgr else "Manager",
                "Manager",
                mp_gross,
                _apply_global_deductions(mp_gross, global_deds),
                0.0,
                0.0,
                0.0,
                net,
                (p.get("manager_pay") or {}).get("note") or "",
                editable,
                str(p.get("_id")),
                {"manager_pay_amount": mp_gross}
            ))

    recs = list(payroll_records_col.find({"month": month}))
    for r in recs:
        emp_id = str(r.get("employee_id") or "")
        if not emp_id or emp_id in used_ids:
            continue

        rec = _recompute_record(emp_id, month) or r
        used_ids.add(emp_id)

        gross = _to_float(rec.get("gross"), 0.0)
        net = _to_float(rec.get("net_pay"), 0.0)
        gross_total += gross
        net_total += net

        rows.append(_overview_row(
            "record",
            emp_id,
            rec.get("employee_name") or "Employee",
            rec.get("role") or "Other",
            gross,
            0.0,
            _to_float(rec.get("deductions_total"), 0.0),
            0.0,
            0.0,
            net,
            rec.get("note") or "",
            True,
            "",
            {
                "basic_salary": rec.get("basic_salary") or 0.0,
                "allowances": rec.get("allowances") or 0.0,
                "paye": rec.get("paye") or 0.0,
            }
        ))

    rows = sorted(rows, key=lambda x: (-(x.get("net") or 0), x.get("name") or ""))
    employee_count = len(rows)

    top = rows[0] if rows else None
    low = rows[-1] if rows else None

    return jsonify(ok=True, branch="All Branches", month=month, month_label=_month_label(month), summary={
        "submitted_total": _money(submitted_total),
        "final_total": _money(gross_total),
        "net_final_total": _money(net_total),
        "employee_count": employee_count,
        "edited_count": edited_count,
        "has_default_tax": has_tax,
    }, top_low={
        "top": {"name": top.get("name"), "role": top.get("role"), "net": top.get("net")} if top else None,
        "low": {"name": low.get("name"), "role": low.get("role"), "net": low.get("net")} if low else None,
    }, rows=rows), 200


# ---------------------------- JSON: branch overview (simple UI) ----------------------------
@hr_payroll_bp.route("/branch/<branch>/overview", methods=["GET"])
def hr_branch_overview(branch):
    month  = request.args.get("month") or ""
    status = request.args.get("status") or "Submitted"

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    branch_key = _norm_branch(branch)
    q = {"payroll_month": month}
    if status and status != "All":
        q["status"] = status

    mgr_by_str, mgr_by_oid = _manager_maps()
    cur = payrolls_col.find(q, {"items": 1, "manager_id": 1, "status": 1, "global_deductions": 1, "manager_pay": 1})

    rows: List[Dict[str, Any]] = []
    used_ids = set()

    submitted_total = 0.0
    gross_total = 0.0
    net_total = 0.0
    edited_count = 0
    has_tax = False

    for p in cur:
        items = p.get("items", []) or []
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)
        inferred = _norm_branch((mgr.get("branch") or "").strip()) if mgr else _infer_branch_from_items(items)

        belongs = (_norm_branch(inferred) == branch_key) or any(
            _norm_branch(it.get("agent_branch") or "") == branch_key for it in items
        )
        if not belongs:
            continue

        global_deds = p.get("global_deductions") or []
        if any((d.get("code") or "") == DEFAULT_TAX_CODE for d in global_deds):
            has_tax = True

        editable = p.get("status") in EDITABLE_HR_STATUSES

        for it in items:
            agent_branch = _norm_branch(it.get("agent_branch") or "")
            if agent_branch != branch_key:
                continue

            agent_id = str(it.get("agent_id") or "")
            if not agent_id:
                continue

            used_ids.add(agent_id)
            ma = _to_float(it.get("manager_amount"), 0.0)
            submitted_total += ma
            if it.get("changed_by_hr"):
                edited_count += 1

            it2 = dict(it)
            it2["deductions"] = _load_deductions(agent_id, month)
            it2["adjustments"] = it2.get("adjustments") or []

            gross = _final_gross_agent(it2)
            net = _net_agent(it2, global_deds)
            global_deds_total = _apply_global_deductions(gross, global_deds)
            other_deds_total = _sum_deductions(it2.get("deductions"))
            adj_add, adj_deduct = _sum_adjustments(it2.get("adjustments"))

            gross_total += gross
            net_total += net

            rows.append(_overview_row(
                "agent",
                agent_id,
                it2.get("agent_name") or "Agent",
                "Agent",
                gross,
                global_deds_total,
                other_deds_total,
                adj_add,
                adj_deduct,
                net,
                it2.get("hr_note") or "",
                editable,
                str(p.get("_id")),
                {
                    "hr_amount": it2.get("hr_amount"),
                    "hr_note": it2.get("hr_note") or "",
                }
            ))

        mp_gross = _manager_pay_final(p)
        if mp_gross > 0:
            manager_id = str(p.get("manager_id") or "")
            used_ids.add(manager_id)

            net = _manager_pay_net(p)
            gross_total += mp_gross
            net_total += net

            rows.append(_overview_row(
                "manager",
                manager_id,
                mgr.get("name") if mgr else "Manager",
                "Manager",
                mp_gross,
                _apply_global_deductions(mp_gross, global_deds),
                0.0,
                0.0,
                0.0,
                net,
                (p.get("manager_pay") or {}).get("note") or "",
                editable,
                str(p.get("_id")),
                {"manager_pay_amount": mp_gross}
            ))

    recs = list(payroll_records_col.find({"month": month}))
    for r in recs:
        emp_id = str(r.get("employee_id") or "")
        if not emp_id or emp_id in used_ids:
            continue
        if _norm_branch(r.get("branch") or "") != branch_key:
            continue

        rec = _recompute_record(emp_id, month) or r
        used_ids.add(emp_id)

        gross = _to_float(rec.get("gross"), 0.0)
        net = _to_float(rec.get("net_pay"), 0.0)
        gross_total += gross
        net_total += net

        rows.append(_overview_row(
            "record",
            emp_id,
            rec.get("employee_name") or "Employee",
            rec.get("role") or "Other",
            gross,
            0.0,
            _to_float(rec.get("deductions_total"), 0.0),
            0.0,
            0.0,
            net,
            rec.get("note") or "",
            True,
            "",
            {
                "basic_salary": rec.get("basic_salary") or 0.0,
                "allowances": rec.get("allowances") or 0.0,
                "paye": rec.get("paye") or 0.0,
            }
        ))

    rows = sorted(rows, key=lambda x: (-(x.get("net") or 0), x.get("name") or ""))
    employee_count = len(rows)

    top = rows[0] if rows else None
    low = rows[-1] if rows else None

    return jsonify(ok=True, branch=branch_key, month=month, month_label=_month_label(month), summary={
        "submitted_total": _money(submitted_total),
        "final_total": _money(gross_total),
        "net_final_total": _money(net_total),
        "employee_count": employee_count,
        "edited_count": edited_count,
        "has_default_tax": has_tax,
    }, top_low={
        "top": {"name": top.get("name"), "role": top.get("role"), "net": top.get("net")} if top else None,
        "low": {"name": low.get("name"), "role": low.get("role"), "net": low.get("net")} if low else None,
    }, rows=rows), 200


# ---------------------------- JSON: payroll details (+ net stats + manager pay + auto tax injection) ----------------------------
@hr_payroll_bp.route("/details/<payroll_id>", methods=["GET"])
def hr_payroll_details(payroll_id):
    p = payrolls_col.find_one({"_id": _oid(payroll_id)}) if _oid(payroll_id) else None
    if not p:
        return jsonify(ok=False, message="Payroll not found."), 404

    mgr_by_str, mgr_by_oid = _manager_maps()
    mgr = _get_manager_from_payroll(p.get("manager_id"), mgr_by_str, mgr_by_oid)

    items = p.get("items", []) or []
    inferred_branch = _norm_branch((mgr.get("branch") or "").strip()) if mgr else _infer_branch_from_items(items)

    # ✅ auto-inject default tax (editable only) if missing
    try:
        month = _normalize_month(p.get("payroll_month") or "")
        eff_tax, _, _ = _get_tax_default(month, inferred_branch)
        injected = _ensure_default_tax_on_payroll(p, eff_tax)
        if injected:
            p = payrolls_col.find_one({"_id": p["_id"]}) or p
    except Exception:
        pass

    global_deds = p.get("global_deductions") or []
    payroll_month = p.get("payroll_month") or ""

    for it in items:
        agent_id = str(it.get("agent_id") or "")
        it["agent_branch"] = _norm_branch(it.get("agent_branch") or "")
        it["manager_amount"] = _to_float(it.get("manager_amount"), 0.0)

        if it.get("hr_amount") is not None and it.get("hr_amount") != "":
            it["hr_amount"] = _to_float(it.get("hr_amount"), None)
        else:
            it["hr_amount"] = None

        if payroll_month and agent_id:
            it["deductions"] = _load_deductions(agent_id, payroll_month)
        else:
            it["deductions"] = it.get("deductions") or []
        it["adjustments"] = it.get("adjustments") or []   # ? NEW

        it["net_final"] = _net_agent(it, global_deds)

    totals = _recompute_totals({"items": items, "global_deductions": global_deds, "manager_pay": p.get("manager_pay")})
    stats  = _compute_stats(items, global_deds)

    mp = p.get("manager_pay") or {}
    manager_pay = {
        "amount": _to_float(mp.get("amount"), 0.0) if (mp.get("amount") not in [None, ""]) else None,
        "note": (mp.get("note") or "").strip(),
        "updated_at": mp.get("updated_at").isoformat() if mp.get("updated_at") else None,
    }

    return jsonify(ok=True, payroll={
        "id": str(p["_id"]),
        "payroll_month": p.get("payroll_month"),
        "month_label": _month_label(p.get("payroll_month")),
        "status": p.get("status", "Draft"),
        "submitted_at": p.get("submitted_at").isoformat() if p.get("submitted_at") else None,
        "updated_at": p.get("updated_at").isoformat() if p.get("updated_at") else None,
        "hr_action_at": p.get("hr_action_at").isoformat() if p.get("hr_action_at") else None,
        "hr_action_by": p.get("hr_action_by"),
        "hr_comment": p.get("hr_comment") or "",
        "manager": {
            "id": str(p.get("manager_id") or ""),
            "name": mgr.get("name") or "Manager",
            "branch": _norm_branch(mgr.get("branch") or inferred_branch or "Unknown"),
            "phone": mgr.get("phone") or "",
            "email": mgr.get("email") or "",
            "image_url": mgr.get("image_url") or "",
        },
        "global_deductions": global_deds,
        "manager_pay": manager_pay,
        "totals": totals,
        "stats": stats,
        "items": items
    }), 200


# ---------------------------- JSON: set global deductions ----------------------------
@hr_payroll_bp.route("/update-global-deductions", methods=["POST"])
def hr_update_global_deductions():
    data = request.get_json(silent=True) or {}
    payroll_id = data.get("payroll_id")
    global_deds = data.get("global_deductions", [])

    if not payroll_id:
        return jsonify(ok=False, message="Missing payroll_id."), 400

    p = payrolls_col.find_one({"_id": _oid(payroll_id)}) if _oid(payroll_id) else None
    if not p:
        return jsonify(ok=False, message="Payroll not found."), 404

    status = p.get("status", "Draft")
    if status not in EDITABLE_HR_STATUSES:
        return jsonify(ok=False, message=f"Cannot edit payroll in status '{status}'."), 409

    clean = []
    for d in (global_deds or []):
        name = (d.get("name") or "").strip() or "Deduction"
        dtype = (d.get("type") or "fixed").lower().strip()
        if dtype not in {"fixed", "percent"}:
            dtype = "fixed"
        value = _to_float(d.get("value"), 0.0)
        if value < 0:
            value = 0.0
        code = (d.get("code") or "").strip() or None
        row = {"name": name, "type": dtype, "value": _money(value)}
        if code:
            row["code"] = code
        clean.append(row)

    now = _now()
    payrolls_col.update_one(
        {"_id": _oid(payroll_id)},
        {"$set": {"global_deductions": clean, "updated_at": now, "status": "Under Review" if status == "Submitted" else status}}
    )

    p2 = payrolls_col.find_one({"_id": _oid(payroll_id)})
    totals = _recompute_totals(p2 or {})
    payrolls_col.update_one({"_id": _oid(payroll_id)}, {"$set": {"totals": totals, "updated_at": now}})

    return jsonify(ok=True, message="Global deductions updated.", totals=totals, global_deductions=clean, status=(p2.get("status") if p2 else None)), 200


# ---------------------------- JSON: update agent HR amount/note + per-agent deductions ----------------------------
@hr_payroll_bp.route("/update-item", methods=["POST"])
def hr_update_item():
    data = request.get_json(silent=True) or {}
    payroll_id = data.get("payroll_id")
    agent_id   = str(data.get("agent_id") or "").strip()

    hr_amount  = data.get("hr_amount", None)
    hr_note    = (data.get("hr_note") or "").strip()
    deductions = data.get("deductions", None)

    if not payroll_id or not agent_id:
        return jsonify(ok=False, message="Missing payroll_id or agent_id."), 400

    p = payrolls_col.find_one({"_id": _oid(payroll_id)}) if _oid(payroll_id) else None
    if not p:
        return jsonify(ok=False, message="Payroll not found."), 404

    status = p.get("status", "Draft")
    if status not in EDITABLE_HR_STATUSES:
        return jsonify(ok=False, message=f"Cannot edit payroll in status '{status}'."), 409

    parsed = None
    if hr_amount is not None and hr_amount != "":
        try:
            parsed = float(hr_amount)
        except Exception:
            return jsonify(ok=False, message="Invalid HR amount."), 400
        if parsed < 0:
            parsed = 0.0

    items = p.get("items", []) or []
    target = next((it for it in items if str(it.get("agent_id")) == agent_id), None)
    if not target:
        return jsonify(ok=False, message="Agent not found in payroll."), 404

    manager_amount = _to_float(target.get("manager_amount"), 0.0)
    changed = bool(parsed is not None and _money(parsed) != _money(manager_amount))

    now = _now()
    set_doc = {
        "items.$.hr_amount": parsed,
        "items.$.hr_note": hr_note,
        "items.$.changed_by_hr": changed,
        "items.$.hr_last_edited_at": now,
        "updated_at": now,
        "status": "Under Review" if status == "Submitted" else status,
    }

    if deductions is not None:
        clean_deds = []
        for d in (deductions or []):
            name = (d.get("name") or "").strip() or "Deduction"
            amt = _to_float(d.get("amount"), 0.0)
            if amt < 0:
                amt = 0.0
            clean_deds.append({"name": name, "amount": _money(amt)})
        set_doc["items.$.deductions"] = clean_deds

    payrolls_col.update_one(
        {"_id": _oid(payroll_id), "items.agent_id": agent_id},
        {"$set": set_doc}
    )

    p2 = payrolls_col.find_one({"_id": _oid(payroll_id)})
    totals = _recompute_totals(p2 or {})
    payrolls_col.update_one({"_id": _oid(payroll_id)}, {"$set": {"totals": totals, "updated_at": now}})

    return jsonify(ok=True, message="Updated.", totals=totals, status=(p2.get("status") if p2 else None)), 200


# ---------------------------- ✅ NEW: Add/Deduct Adjustment Per Agent ----------------------------
@hr_payroll_bp.route("/adjust-agent", methods=["POST"])
def hr_adjust_agent():
    data = request.get_json(silent=True) or {}
    payroll_id = data.get("payroll_id")
    agent_id   = str(data.get("agent_id") or "").strip()
    action     = (data.get("action") or "").strip().lower()
    amount     = data.get("amount", None)
    reason     = (data.get("reason") or "").strip()

    if not payroll_id or not agent_id:
        return jsonify(ok=False, message="Missing payroll_id or agent_id."), 400
    if action not in {"add", "deduct"}:
        return jsonify(ok=False, message="Invalid action. Use add/deduct."), 400
    if not reason:
        return jsonify(ok=False, message="Reason is required."), 400

    try:
        amt = float(amount)
    except Exception:
        return jsonify(ok=False, message="Invalid amount."), 400
    if amt <= 0:
        return jsonify(ok=False, message="Amount must be greater than 0."), 400

    p = payrolls_col.find_one({"_id": _oid(payroll_id)}) if _oid(payroll_id) else None
    if not p:
        return jsonify(ok=False, message="Payroll not found."), 404

    status = p.get("status", "Draft")
    if status not in EDITABLE_HR_STATUSES:
        return jsonify(ok=False, message=f"Cannot edit payroll in status '{status}'."), 409

    now = _now()
    adj = {
        "type": action,
        "amount": _money(amt),
        "reason": reason,
        "at": now,
        "by": "system"
    }

    payrolls_col.update_one(
        {"_id": _oid(payroll_id), "items.agent_id": agent_id},
        {"$push": {"items.$.adjustments": adj},
         "$set": {
             "items.$.changed_by_hr": True,
             "items.$.hr_last_edited_at": now,
             "updated_at": now,
             "status": "Under Review" if status == "Submitted" else status
         }}
    )

    p2 = payrolls_col.find_one({"_id": _oid(payroll_id)})
    totals = _recompute_totals(p2 or {})
    payrolls_col.update_one({"_id": _oid(payroll_id)}, {"$set": {"totals": totals, "updated_at": now}})

    agent_net = 0.0
    if p2:
        global_deds = p2.get("global_deductions") or []
        for it in (p2.get("items") or []):
            if str(it.get("agent_id")) == agent_id:
                agent_net = _net_agent(it, global_deds)
                break

    return jsonify(ok=True, message="Adjustment saved.", totals=totals, agent_net=agent_net), 200


# ---------------------------- JSON: update Manager Pay (HR only) ----------------------------
@hr_payroll_bp.route("/update-manager-pay", methods=["POST"])
def hr_update_manager_pay():
    data = request.get_json(silent=True) or {}
    payroll_id = data.get("payroll_id")
    amount = data.get("amount", None)
    note = (data.get("note") or "").strip()

    if not payroll_id:
        return jsonify(ok=False, message="Missing payroll_id."), 400

    p = payrolls_col.find_one({"_id": _oid(payroll_id)}) if _oid(payroll_id) else None
    if not p:
        return jsonify(ok=False, message="Payroll not found."), 404

    status = p.get("status", "Draft")
    if status not in EDITABLE_HR_STATUSES:
        return jsonify(ok=False, message=f"Cannot edit payroll in status '{status}'."), 409

    parsed = None
    if amount is not None and amount != "":
        try:
            parsed = float(amount)
        except Exception:
            return jsonify(ok=False, message="Invalid amount."), 400
        if parsed < 0:
            parsed = 0.0

    now = _now()
    payrolls_col.update_one(
        {"_id": _oid(payroll_id)},
        {"$set": {
            "manager_pay": {"amount": parsed, "note": note, "updated_at": now},
            "updated_at": now,
            "status": "Under Review" if status == "Submitted" else status,
        }}
    )

    p2 = payrolls_col.find_one({"_id": _oid(payroll_id)})
    totals = _recompute_totals(p2 or {})
    payrolls_col.update_one({"_id": _oid(payroll_id)}, {"$set": {"totals": totals, "updated_at": now}})

    return jsonify(ok=True, message="Manager pay updated.", totals=totals), 200


# ---------------------------- JSON: approve / reject payroll ----------------------------
@hr_payroll_bp.route("/decision", methods=["POST"])
def hr_payroll_decision():
    data = request.get_json(silent=True) or {}
    payroll_id = data.get("payroll_id")
    action     = (data.get("action") or "").strip().lower()
    comment    = (data.get("comment") or "").strip()

    if not payroll_id or action not in {"approve", "reject"}:
        return jsonify(ok=False, message="Invalid request."), 400

    p = payrolls_col.find_one({"_id": _oid(payroll_id)}) if _oid(payroll_id) else None
    if not p:
        return jsonify(ok=False, message="Payroll not found."), 404

    current = p.get("status", "Draft")
    if current not in DECISION_STATUSES:
        return jsonify(ok=False, message=f"Cannot approve/reject payroll in status '{current}'."), 409

    now = _now()
    new_status = "Approved" if action == "approve" else "Rejected"

    payrolls_col.update_one(
        {"_id": _oid(payroll_id)},
        {"$set": {
            "status": new_status,
            "hr_action_at": now,
            "hr_action_by": "system",
            "hr_comment": comment,
            "updated_at": now
        },
        "$push": {"activity_log": {
            "action": new_status.lower(),
            "by": "system",
            "at": now,
            "note": comment or f"HR {new_status}"
        }}}
    )

    return jsonify(ok=True, message=f"Payroll {new_status}.", status=new_status), 200


# ---------------------------- JSON: branch stats (NET stats) ----------------------------
@hr_payroll_bp.route("/branch/<branch>/stats", methods=["GET"])
def hr_branch_stats(branch):
    month = request.args.get("month") or ""
    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    branch_key = _norm_branch(branch)

    q = {"payroll_month": month}
    cur = payrolls_col.find(q, {"items": 1, "global_deductions": 1, "manager_id": 1})

    all_items = []
    gd = []

    mgr_by_str, mgr_by_oid = _manager_maps()

    for p in cur:
        items = p.get("items", []) or []
        gd = p.get("global_deductions") or gd

        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)
        b = _norm_branch((mgr.get("branch") or "").strip()) if mgr else _infer_branch_from_items(items)

        if _norm_branch(b) != branch_key:
            items = [it for it in items if _norm_branch(it.get("agent_branch") or "") == branch_key]

        for it in items:
            emp_id = str(it.get("agent_id") or "")
            it["agent_branch"] = _norm_branch(it.get("agent_branch") or "")
            if emp_id:
                it["deductions"] = _load_deductions(emp_id, month)
            else:
                it["deductions"] = it.get("deductions") or []
            it["adjustments"] = it.get("adjustments") or []

        all_items.extend(items)

    stats = _compute_stats(all_items, gd)
    return jsonify(ok=True, branch=branch_key, month=month, stats=stats), 200


# ---------------------------- EXPORT: Branch PDF ----------------------------
@hr_payroll_bp.route("/branch/<branch>/export/pdf", methods=["GET"])
def hr_branch_export_pdf(branch):
    month  = request.args.get("month") or ""
    status = request.args.get("status") or "Submitted"

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    branch_key = _norm_branch(branch)

    q = {"payroll_month": month}
    if status and status != "All":
        q["status"] = status

    mgr_by_str, mgr_by_oid = _manager_maps()
    cur = list(payrolls_col.find(q, {"items": 1, "manager_id": 1, "status": 1, "global_deductions": 1, "manager_pay": 1, "totals": 1}).sort("updated_at", -1))

    rows = []
    for p in cur:
        items = p.get("items", []) or []
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)
        inferred = _norm_branch((mgr.get("branch") or "").strip()) if mgr else _infer_branch_from_items(items)

        belongs = (_norm_branch(inferred) == branch_key) or any(_norm_branch(it.get("agent_branch") or "") == branch_key for it in items)
        if not belongs:
            continue

        global_deds = p.get("global_deductions") or []
        for it in items:
            if _norm_branch(it.get("agent_branch") or "") != branch_key:
                continue
            it2 = dict(it)
            it2["manager_amount"] = _to_float(it2.get("manager_amount"), 0.0)
            if it2.get("hr_amount") in [None, ""]:
                it2["hr_amount"] = None
            else:
                it2["hr_amount"] = _to_float(it2.get("hr_amount"), None)
            agent_id = str(it2.get("agent_id") or "")
            if agent_id:
                it2["deductions"] = _load_deductions(agent_id, month)
            else:
                agent_id = str(it2.get("agent_id") or "")
            if agent_id:
                it2["deductions"] = _load_deductions(agent_id, month)
            else:
                it2["deductions"] = it2.get("deductions") or []
            it2["adjustments"] = it2.get("adjustments") or []

            gross = _final_gross_agent(it2)
            net = _net_agent(it2, global_deds)

            rows.append({
                "name": it2.get("agent_name") or "Agent",
                "role": "Agent",
                "branch": branch_key,
                "submitted": it2["manager_amount"],
                "gross": gross,
                "net": net,
                "note": (it2.get("hr_note") or "").strip()
            })

        mp_gross = _manager_pay_final(p)
        if mp_gross > 0:
            rows.append({
                "name": (mgr.get("name") if mgr else "Manager"),
                "role": "Manager",
                "branch": branch_key,
                "submitted": 0.0,
                "gross": mp_gross,
                "net": _manager_pay_net(p),
                "note": ((p.get("manager_pay") or {}).get("note") or "").strip()
            })

    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.lib.utils import ImageReader
        import urllib.request

        buff = BytesIO()
        doc = SimpleDocTemplate(buff, pagesize=landscape(A4), leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)

        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(f"<b>Smart Living — Branch Payroll Report</b>", styles["Title"]))
        story.append(Paragraph(f"Branch: <b>{branch_key}</b> &nbsp;&nbsp; Month: <b>{_month_label(month)}</b> &nbsp;&nbsp; Status: <b>{status}</b>", styles["Normal"]))
        story.append(Spacer(1, 10))

        data = [["#", "Name", "Role", "Submitted (₵)", "Final Gross (₵)", "Net Final (₵)", "Note"]]
        for i, r in enumerate(rows, start=1):
            data.append([
                str(i),
                r["name"],
                r["role"],
                f"{_money(r['submitted']):,.2f}",
                f"{_money(r['gross']):,.2f}",
                f"{_money(r['net']):,.2f}",
                (r["note"] or "")
            ])

        tbl = Table(data, colWidths=[0.45*inch, 2.4*inch, 1.0*inch, 1.35*inch, 1.35*inch, 1.35*inch, 3.1*inch])
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.whitesmoke),
            ("TEXTCOLOR", (0,0), (-1,0), colors.black),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,0), 10),
            ("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("FONTSIZE", (0,1), (-1,-1), 9),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.Color(0.98,0.99,1)]),
        ]))
        story.append(tbl)

        def _watermark(canv, _doc):
            try:
                with urllib.request.urlopen(WATERMARK_URL, timeout=5) as resp:
                    img_bytes = resp.read()
                img = ImageReader(BytesIO(img_bytes))
                w, h = landscape(A4)
                canv.saveState()
                canv.setFillAlpha(0.08)
                iw, ih = 520, 520
                canv.drawImage(img, (w-iw)/2, (h-ih)/2, width=iw, height=ih, mask='auto')
                canv.restoreState()
            except Exception:
                pass

        doc.build(story, onFirstPage=_watermark, onLaterPages=_watermark)
        buff.seek(0)

        filename = f"payroll_{branch_key}_{month}_{status}.pdf".replace(" ", "_")
        return send_file(buff, mimetype="application/pdf", as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify(ok=False, message=f"PDF export failed: {e}"), 500


# ---------------------------- EXPORT: Branch Excel ----------------------------
@hr_payroll_bp.route("/branch/<branch>/export/xlsx", methods=["GET"])
def hr_branch_export_xlsx(branch):
    month  = request.args.get("month") or ""
    status = request.args.get("status") or "Submitted"

    try:
        month = _normalize_month(month)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 400

    branch_key = _norm_branch(branch)

    q = {"payroll_month": month}
    if status and status != "All":
        q["status"] = status

    mgr_by_str, mgr_by_oid = _manager_maps()
    cur = list(payrolls_col.find(q, {"items": 1, "manager_id": 1, "status": 1, "global_deductions": 1, "manager_pay": 1}).sort("updated_at", -1))

    rows = []
    for p in cur:
        items = p.get("items", []) or []
        mid = p.get("manager_id")
        mgr = _get_manager_from_payroll(mid, mgr_by_str, mgr_by_oid)
        inferred = _norm_branch((mgr.get("branch") or "").strip()) if mgr else _infer_branch_from_items(items)

        belongs = (_norm_branch(inferred) == branch_key) or any(_norm_branch(it.get("agent_branch") or "") == branch_key for it in items)
        if not belongs:
            continue

        global_deds = p.get("global_deductions") or []

        for it in items:
            if _norm_branch(it.get("agent_branch") or "") != branch_key:
                continue
            it2 = dict(it)
            it2["manager_amount"] = _to_float(it2.get("manager_amount"), 0.0)
            if it2.get("hr_amount") in [None, ""]:
                it2["hr_amount"] = None
            else:
                it2["hr_amount"] = _to_float(it2.get("hr_amount"), None)
            agent_id = str(it2.get("agent_id") or "")
            if agent_id:
                it2["deductions"] = _load_deductions(agent_id, month)
            else:
                it2["deductions"] = it2.get("deductions") or []
            it2["adjustments"] = it2.get("adjustments") or []

            gross = _final_gross_agent(it2)
            net = _net_agent(it2, global_deds)

            rows.append([
                it2.get("agent_name") or "Agent",
                "Agent",
                branch_key,
                it2["manager_amount"],
                gross,
                net,
                (it2.get("hr_note") or "").strip(),
            ])

        mp_gross = _manager_pay_final(p)
        if mp_gross > 0:
            rows.append([
                (mgr.get("name") if mgr else "Manager"),
                "Manager",
                branch_key,
                0.0,
                mp_gross,
                _manager_pay_net(p),
                ((p.get("manager_pay") or {}).get("note") or "").strip(),
            ])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = f"{branch_key[:28]}"

        headers = ["Name", "Role", "Branch", "Submitted", "Final Gross", "Net Final", "Note"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")

        for r in rows:
            ws.append(r)

        widths = [28, 12, 16, 14, 14, 14, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
            for cell in row:
                cell.number_format = '#,##0.00'

        buff = BytesIO()
        wb.save(buff)
        buff.seek(0)

        filename = f"payroll_{branch_key}_{month}_{status}.xlsx".replace(" ", "_")
        return send_file(buff, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify(ok=False, message=f"Excel export failed: {e}"), 500
